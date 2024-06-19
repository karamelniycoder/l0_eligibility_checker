from openpyxl.styles import Color, PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
from datetime import datetime
from loguru import logger
from time import sleep
import os


class Excel:
    def __init__(self, total_len: int, name: str):
        if not os.path.isdir('results'): os.mkdir('results')
        self.name = name

        workbook = Workbook()
        sheet = workbook.active
        self.file_name = f'{name}_{total_len}accs_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'

        sheet['A1'] = 'EVM Address'
        sheet['B1'] = 'Status'
        sheet['C1'] = 'Tokens'
        sheet.column_dimensions['A'].width = 47
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 18

        for cell in sheet._cells:
            sheet.cell(cell[0], cell[1]).font = Font(bold=True)
            sheet.cell(cell[0], cell[1]).alignment = Alignment(horizontal='center')
            sheet.cell(cell[0], cell[1]).border = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))

        workbook.save('results/'+self.file_name)


    def edit_table(self, wallet_data: list):
        while True:
            try:
                workbook = load_workbook('results/'+self.file_name)
                sheet = workbook.active

                sheet.append([*wallet_data])

                for row_cells in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row):
                    for cell in row_cells:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))
                        if cell.column == 2 and cell.value:
                            if cell.value == "Eligible": rgb_color = "00FF00"
                            else: rgb_color = "FF0000"
                            cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb=rgb_color))

                workbook.save('results/'+self.file_name)
                return True
            except PermissionError:
                logger.warning(f'Excel | Cant save excel file, close it!')
                sleep(3)
            except Exception as err:
                logger.critical(f'Excel | Cant save excel file: {err} | {wallet_data[0]}')
                return False